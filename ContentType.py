import argparse
import openpyxl
import requests
import time
import json
import sys
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font
from requests.exceptions import RequestException

def parse_args():
    parser = argparse.ArgumentParser(description='Process article data from API and update Excel file.')
    parser.add_argument('excel_path', help='Path to Excel file (Windows or Unix format)')
    parser.add_argument('--sheet-name', default='ids', help='Excel sheet name (default: ids)')
    parser.add_argument('--type-id', type=int, default=0, help='Type ID parameter (default: 0)')
    parser.add_argument('--interface-id', type=int, default=3, help='Interface ID parameter (default: 3)')
    parser.add_argument('--title-key', default='title', help='JSON key for title (default: title)')
    parser.add_argument('--body-key', default='body', help='JSON key for body (default: body)')
    return parser.parse_args()

def process_excel(sheet, args):
    headers = [cell.value for cell in sheet[1]]
    
    # Add missing columns if needed
    new_headers = headers.copy()
    columns_added = False
    
    if 'Title' not in new_headers:
        new_headers.append('Title')
        columns_added = True
    if 'Body' not in new_headers:
        new_headers.append('Body')
        columns_added = True
    if 'Error' not in new_headers:
        new_headers.append('Error')
        columns_added = True
    # Define font style for the header (bold)
    header_font = Font(bold=True)
    
    if columns_added:
        for idx, header in enumerate(new_headers, 1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = header
            cell.font = header_font
        print("DEBUG: Added Title/Body columns to sheet")

    # Get column indices
    id_col_idx = headers.index('ID') + 1  # 1-based index
    title_col_idx = new_headers.index('Title') + 1
    body_col_idx = new_headers.index('Body') + 1
    error_col_idx = new_headers.index('Error') + 1

    return id_col_idx, title_col_idx, body_col_idx, error_col_idx

def update_error(errors, id_value, row_num, url):
    entry = {
        "id": id_value,
        "row": row_num,
        "url": url
    }
    errors.append(entry)
    return errors

def get_error_url(id):
    # Page url
    return f"<URL>?id={id}"

def main():
    args = parse_args()
    
    try:
        wb = openpyxl.load_workbook(args.excel_path)
    except FileNotFoundError:
        print(f"ERROR: File not found: {args.excel_path}")
        sys.exit(1)
    except InvalidFileException:
        print(f"ERROR: Invalid Excel file: {args.excel_path}")
        sys.exit(1)

    if args.sheet_name not in wb.sheetnames:
        print("ERROR: {args.sheet_name} not found in workbook")
        sys.exit(1)
        
    sheet = wb[args.sheet_name]
    print(f"DEBUG: Opened sheet: {sheet.title}")

    try:
        id_col_idx, title_col_idx, body_col_idx, error_col_idx = process_excel(sheet, args)
    except ValueError:
        print("ERROR: ID column not found in {args.sheet_name}")
        sys.exit(1)

    # Collect valid IDs with their row numbers
    id_rows = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            id_value = int(row[id_col_idx - 1].value)  # Adjust for 0-based index
            if 100 <= id_value <= 100000:
                id_rows.append((id_value, row_idx))
                # print(f"DEBUG: Found valid ID {id_value} at row {row_idx}")
        except (ValueError, TypeError):
            continue

    if not id_rows:
        print("DEBUG: No valid IDs found in the sheet")
        return

    #Enpoint URL
    base_url = ""
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    error_ids = []
    for id_value, row_num in id_rows:
        # print(f"\nDEBUG: Processing ID {id_value} (row {row_num})")
        params = {
            'questionId': id_value,
            'typeId': args.type_id,
            'id': id_value,
            'interfaceId': args.interface_id,
            '_': int(time.time() * 1000)
        }
        
        try:
            response = requests.get(base_url, headers=headers, params=params)
            # print(f"DEBUG: Called URL: {response.url}")
            error_url = get_error_url(id_value)
            if response.status_code != 200:
                # print(f"ERROR: {response.url} returned status {response.status_code}")
                update_error(error_ids, id_value, row_num, response.url)
                sheet.cell(row=row_num, column=error_col_idx, value=error_url)
                continue
                
            content_type = response.headers.get('Content-Type', '')
            if 'application/json' not in content_type:
                print(f"DEBUG: Non-JSON response from {response.url}")
                continue
                
            try:
                data = response.json()
                length = len(data)
                if(length > 1):
                    print(f"DEBUG: Response has {length} (should be 1) Help & Support Article sets {response.url}")
            except json.JSONDecodeError:
                print(f"DEBUG: Invalid JSON from {response.url}")
                sheet.cell(row=row_num, column=error_col_idx, value=error_url)
                update_error(error_ids, id_value, row_num, response.url)
                continue

            title = data[0].get(args.title_key, '')
            body = data[0].get(args.body_key, '')
            # print(f"DEBUG: Retrieved title: {title[:50]}... | body: {body[:50]}...")
            # Update Excel cells
            sheet.cell(row=row_num, column=title_col_idx, value=title)
            sheet.cell(row=row_num, column=body_col_idx, value=body)
            
            # Save after each update
            wb.save(args.excel_path)
            print(f"DEBUG: Updated row {row_num} with ID {id_value} and saved the file")
            
        except RequestException as e:
            print(f"ERROR: Request failed for ID {id_value}: {str(e)}")
            sheet.cell(row=row_num, column=error_col_idx, value=error_url)
            update_error(error_ids, id_value, row_num, response.url)
            continue
            
    print(f"ERROR: ERRORS")
    for error in error_ids:
        print(f"ERROR: Invalid URL {get_error_url(error['id'])} \n\tExcel row: {error['row']} \n\tEndpoint: {error['url']}")
    
    print("\nDEBUG: Processing completed")

if __name__ == "__main__":
    main()